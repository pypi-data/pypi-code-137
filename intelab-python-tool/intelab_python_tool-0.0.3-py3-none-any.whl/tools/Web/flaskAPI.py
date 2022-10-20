from datetime import datetime
import os
from intelab_python_tool.tools.log import log, log_init
from flask import Flask, jsonify, request
from flask.views import MethodView
from apscheduler.schedulers.blocking import BlockingScheduler
import requests
import openpyxl

"""
使用方法如下：
    
    from intelab_python_tool.tools.Web.flaskAPI import Monitor
    Mt = Monitor(scheduler_list)
    Mt.run()
    
    scheduler_list: list
    such as :  [
                [service_name, monitor_log_path, monitor_type, scheduler_period_type, 
                scheduler_period, dingtalk_function
                restart_flag, args],
                实例：
                ['event_verifier', '/var/log/utilization_producer/log/timing_utilization_producer.info.log.',
                'l', 'm', '3', 0, 1],
                ]
    
    service_name: 如果监控日志，则是重启服务时的服务名；如果是接口测试，则是url
    monitor_log_path: 如果监控日志，则是监控的日志文件的绝对路径；如果是接口测试，则是访问接口的方式get or post
    monitor_type: 'l'表示监控日志, 'a'表示监控接口
    scheduler_period_type: 定时任务的类型，'h'表示按小时，'m'表示按分钟
    scheduler_period:  定时任务的时长， 多少小时/多少分钟
    dingtalk_function: 发送报警的函数 为0表示无效
    restart_flag: 是否需要重启服务
    args: 接口测试时需要传入的参数
    
"""

TIMEZONE = 'Asia/Shanghai'


class Monitor:

    def __init__(self, scheduler_list: dict):
        # 初始化输入参数
        self.service_list = self.scheduler_init(scheduler_list)
        # 初始化本地文件储存相应信息
        self.create_local_excel()

        self.time = datetime.now().strftime("%Y-%m-%d")
        self.hour = datetime.now().hour

    def run(self, logs_path, host='0.0.0.0', port=5000):
        log_init('health_monitor', debug=False, log_path=logs_path)

        self.api_run(host=host, port=port)

        self.start_scheduler()

    @staticmethod
    def api_run(host, port):
        app = Flask(__name__)

        @app.route('/api/ai/service_monitor', methods=['GET'])
        def get_service_health():
            record_path = request.args.get('record_path')
            wb = openpyxl.load_workbook(record_path)
            ws = wb.active
            error_times = ws['A1'].value
            wb.close()

            if error_times >= 3:
                response = {'msg': '算法服务重启三次仍无法正常运行，请相关运维人员检查服务配置或网络', 'code': 400,
                            'data': error_times}
            else:
                response = {'msg': '服务正常', 'code': 200, 'data': error_times}
            return jsonify(response)

        # app.config['env'] = 'development'
        app.run(host=host, port=port)

    def monitor_log_run(self, service):

        log.info(f"本次执行健康检测的服务为{service.name}")

        wb = openpyxl.load_workbook('monitor_data.xlsx')
        ws = wb.actice
        old_data_bytes = ws['A1'].value
        check_times = ws['B1'].value
        error_times = ws['C1'].value

        new_data_bytes = os.stat(f"{service.log_path}" + f"{self.time}").st_size
        log.info('主监控日志文件变化前大小为{}，变化后大小为{}'.format(old_data_bytes, new_data_bytes))

        if new_data_bytes != old_data_bytes:
            check_times = 0
        else:
            check_times += 1

        log.info(f'主监控日志check_times为{check_times}')
        if check_times > 3:
            log.info(f'check_times大于3，重启服务')
            self.restart_service(service.name, service.dingtalk_function, service.restart_flag)
            check_times = 0
            error_times += 1
        else:
            error_times = 0

        hour_str = f"{self.hour}" if self.hour > 9 else f"0{self.hour}"
        log_check = os.system(f"cat {service.log_path}" + f"{self.time} | grep {self.time}T{hour_str}")
        log.info(f"执行shell: cat {service.log_path}" + f"{self.time} | grep {self.time}T{hour_str}")

        if log_check != 0:
            log.info(f"{self.time}T{hour_str} 无法在日志文件中查询到，请及时查看处理")
            self.send_alarm2developer(f"{self.time}T{hour_str} 无法在日志文件中查询到，请及时查看处理", service.dingtalk_function)
            self.restart_service(service.name, service.dingtalk_function, service.restart_flag)
            check_times = 0

        ws['A1'].value = new_data_bytes
        ws['B1'].value = check_times
        ws['C1'].value = error_times
        wb.close()

    def restart_service(self, service_name, ding_talk_function, restart_flag) -> None:
        if not restart_flag:
            return
        log.info(f"正在重启{service_name}服务")
        res = os.system(f"supervisorctl restart {service_name}")
        if res == 0:
            log.info(f"重启{service_name}完成")
        else:
            self.send_alarm2developer(f"重启{service_name}失败，请开发人员介入查看", ding_talk_function=ding_talk_function)

    @staticmethod
    def send_alarm2developer(msg: str, ding_talk_function) -> None:
        if isinstance(ding_talk_function, int):
            return
        log.info('警告已发送至开发者')
        ding_talk_function(msg)

    def create_local_excel(self) -> None:
        for service in self.service_list:
            if service.monitor_type == 'l':
                wb = openpyxl.Workbook()
                ws = wb.active
                ws['A1'] = 0    # 储存监控日志大小
                ws['B1'] = 0    # 储存check_times
                ws['C1'] = 0    # 储存error_times
                wb.save(f'{service.name}_monitor.xlsx')
                wb.close()

    def test_request(self, service):
        url = service.name
        request_type = service.log_path
        data = service.args

        if request_type == 'get':
            res = requests.get(url) if not data else requests.get(url, data=data)
        else:
            res = requests.post(url) if not data else requests.post(url, data=data)

        if res.status_code == 200:
            return True
        else:
            msg = f'{service.name}接口调用失败，请及时查看'
            self.send_alarm2developer(msg, service.dingtalk_function)

    def start_scheduler(self) -> None:

        scheduler = BlockingScheduler(
            {'apscheduler.timezone': TIMEZONE}
        )
        for service in self.service_list:
            if service.monitor_type == 'l':
                if service.scheduler_period_type == 'h':
                    scheduler.add_job(self.monitor_log_run(service), 'interval', hours=service.scheduler_period)
                else:
                    scheduler.add_job(self.monitor_log_run(service), 'interval', minutes=service.scheduler_period)
            elif service.monitor_type == 'a':
                if service.scheduler_period_type == 'h':
                    scheduler.add_job(self.test_request(service), 'interval', hours=service.scheduler_period)
                else:
                    scheduler.add_job(self.test_request(service), 'interval', minutes=service.scheduler_period)
            else:
                log.info(f"{service.name}服务监控类型不存在，服务监控类型目前仅支持l,a两种")
        scheduler.start()

    @staticmethod
    def scheduler_init(scheduler_list):
        return [Service(_) for _ in scheduler_list]


class ServiceHealthApi(MethodView):

    def get(self, record_path: str):

        wb = openpyxl.load_workbook(record_path)
        ws = wb.actice
        error_times = ws['C1']
        wb.close()

        if error_times >= 3:
            response = {'msg': '算法服务重启三次仍无法正常运行，请相关运维人员检查服务配置或网络', 'code': 400,
                        'data': error_times}
        else:
            response = {'msg': '服务正常', 'code': 200, 'data': error_times}
        return jsonify(response)


class Service:

    def __init__(self, msg):
        self.name = msg[0]
        self.log_path = msg[1]
        self.monitor_type = msg[2]
        self.scheduler_period_type = msg[3]
        self.scheduler_period = msg[4]
        self.dingtalk_function = msg[5]
        self.restart_flag = msg[6]
        self.args = msg[7] if msg[7] else None


if __name__ == '__main__':
    app = Flask(__name__)

    @app.route('/api/ai/service_monitor', methods=['GET'])
    def get_service_health():
        record_path = request.args.get('record_path')
        wb = openpyxl.load_workbook(record_path)
        ws = wb.active
        error_times = ws['A1'].value
        wb.close()

        if error_times >= 3:
            response = {'msg': '算法服务重启三次仍无法正常运行，请相关运维人员检查服务配置或网络', 'code': 400,
                        'data': error_times}
        else:
            response = {'msg': '服务正常', 'code': 200, 'data': error_times}
        return jsonify(response)

    app.config['env'] = 'development'
    app.run(host='0.0.0.0', port=5001)
